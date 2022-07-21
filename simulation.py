# -*- coding: utf-8 -*-
"""
Created on Thu Jul 21 15:54:10 2022

@author: wyx
"""
import numpy as np
import math
import pandas as pd
import openpyxl
# ------------------------------------------------------------------------------
# Datasets simulation
# ------------------------------------------------------------------------------
n = 300
m = 1
num = 2
sigma_err = 0.5
no = []
for i in range(n):
    for j in range(m):
        no.append(i+1)

z1 = np.zeros((n*m,n))
for i in range(n*m):
    for j in range(1,n+1):
        if no[i]==j:
            z1[i,j-1]=1
         
betaR1_simu = np.array((-1.6,0.6)).reshape(-1,1)
betaR2_simu = np.array((-1.0,0.3)).reshape(-1,1)
betaT_simu  = np.array((1.2,-0.4)).reshape(-1,1)

mu_simu = np.array([0, 0, 0])
theta1_simu = 1
theta2_simu = 1
theta3_simu = 1
ro1_simu = 1
ro2_simu = -0.5
ro3_simu = -0.5
cov_simu = np.array([[theta1_simu,ro1_simu*math.sqrt(theta1_simu*theta2_simu),
                 ro2_simu*math.sqrt(theta1_simu*theta3_simu)],
               [ro1_simu*math.sqrt(theta1_simu*theta2_simu),theta2_simu,ro3_simu*math.sqrt(theta2_simu*theta3_simu)],
               [ro2_simu*math.sqrt(theta1_simu*theta3_simu),ro3_simu*math.sqrt(theta2_simu*theta3_simu),theta3_simu]])

b_simu = np.random.multivariate_normal(mu_simu, cov_simu, n) 
U1_simu = b_simu[:,0:1]
U2_simu = b_simu[:,1:2]
V_simu = b_simu[:,2:3]

x = np.random.rand(n*m).reshape(-1, 1)
TMB = np.random.normal(4.0, 1.0, (n*m, 1))
err = np.random.normal(0, sigma_err, (n*m, 1))
TMB_E = TMB + err
X_simu = np.column_stack((x,TMB))

yetaR1_simu = np.dot(X_simu, betaR1_simu) + np.dot(z1,U1_simu)
yetaR2_simu = np.dot(X_simu, betaR2_simu) + np.dot(z1,U2_simu)
yetaT_simu = np.dot(X_simu, betaT_simu) + np.dot(z1,V_simu)
  
P0 = 1 / (1 + np.exp(yetaR1_simu) + np.exp(yetaR2_simu))
P1 = np.exp(yetaR1_simu) / (1 + np.exp(yetaR1_simu) + np.exp(yetaR2_simu))
rand = np.random.rand(n*m)
R_simu = np.zeros((n*m, 1))

for i in range(n*m):
    if rand[i] > (P0[i]+P1[i]):
        R_simu[i] = 2
    if rand[i] <= (P0[i]+P1[i]) and rand[i] > P0[i]:
        R_simu[i] = 1

lamda = 1/50
T_simu = np.zeros((n*m, 1))
C_simu = np.random.uniform(5, 50, n*m)
for i in range(n*m):
    T_simu[i] = - (1/lamda) * np.log(np.random.rand(1)) / (np.exp(yetaT_simu[i]))
delta = np.zeros((n*m, 1))
T = np.zeros((n*m, 1))
for i in range(n*m):
    delta[i] = 1 if T_simu[i] <= C_simu[i] else 0
    T[i] = T_simu[i] if delta[i] == 1 else C_simu[i]

data = np.column_stack((no,R_simu,T,delta,X_simu,TMB_E))

path = 'D:/WYX/data/simu_data.xlsx'
wb = openpyxl.Workbook()
sh = wb.active
for col_index, col_item in enumerate(['no','R_simu','T','delta','X','TMB','TMB_error']):
    sh.cell(1,col_index+1,value=col_item)
for row_index, row_item in enumerate(data):
    for col_index, col_item in enumerate(row_item):
        sh.cell(row=row_index+2,column= col_index+1,value=col_item)
wb.save(path)
